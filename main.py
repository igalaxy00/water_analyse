from datetime import datetime

import ee
import os

from ee import Geometry
from openpyxl import Workbook, load_workbook
from Py6S import *
import math
from PIL import Image
import rasterio
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
ee.Authenticate()

ee.Initialize()
SixS.test()
# Press the green button in the gutter to run the script.
if __name__ == '__main__':

    iNames = {}

    lon = 45.4954
    lat = 33.5517
    aoi = ee.Geometry.Rectangle([lon, lat, lon + 0.1, lat + 0.1])
    nTemp = 8
    collection = ee.ImageCollection('LANDSAT/LC08/C01/T1_SR') \
        .filterBounds(aoi) \
        .filterDate('2022-01-01', '2022-12-31')


    def maskClouds(image):
        qa = image.select('pixel_qa')
        cloudMask = qa.bitwiseAnd(1 << 5).eq(0)
        return image.updateMask(cloudMask)
    maskedCollection = collection.map(maskClouds)
    def extractLakes(image):
        water = image.normalizedDifference(['B3', 'B5']).gt(0.1)
        return water

    lakes = maskedCollection.map(extractLakes)

    lakeImages = lakes.filter(ee.Filter.gt('nd', 0))

    output_folder = 'path/to/output/folder'
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    image_list = lakeImages.toList(lakeImages.size())
    for i in range(lakeImages.size().getInfo()):
        image = ee.Image(image_list.get(i))
        image_id = image.id().getInfo()
        image_path = os.path.join(output_folder, image_id + '.tif')

        spectra = image.select(['B2', 'B3', 'B4', 'B5', 'B6', 'B7']).sample(aoi)
        iNames[image_id] = spectra

        task = ee.batch.Export.image.toDrive(image, description=image_id, fileNamePrefix=image_id, folder=output_folder,
                                             scale=30)
        task.start()
        print(f'Started exporting {image_id} to {image_path}')
    print('Снимки экспортированы.')

#neiro.start('path/to/output/folder')
#сохраняет в папку "images/roi"

avtemp = {}


def getTemperature(image):
    temp = image.select('B10').multiply(0.1).subtract(273.15)  # конвертация кельвинов в градусы Цельсия
    temp = temp.reduceRegion(
        reducer=ee.Reducer.mean(),
        geometry=aoi,
        scale=30,
        maxPixels=1e9
    )
    return temp

def toa_to_rad(bandname):

    solar_angle_correction = math.cos(math.radians(bandname.index()))
    # Earth-Sun distance (from day of year)
    d = 1 - 0.01672 * math.cos(0.9856 * (-4))# http://physics.stackexchange.com/questions/177949/earth-sun-distance-on-a-given-day-of-the-year
    # conversion factor
    multiplier = solar_angle_correction/(math.pi*d**2)
    # at-sensor radiance
    rad = d*multiplier
    return rad
def surface_reflectance(bandname):
    """
    Calculate surface reflectance from at-sensor radiance given waveband name
    """
    # run 6S for this waveband
    s = SixS()
    s.wavelength = toa_to_rad(bandname)
    s.run()
    # extract 6S outputs
    Edir = s.outputs.direct_solar_irradiance  # direct solar irradiance
    Edif = s.outputs.diffuse_solar_irradiance  # diffuse solar irradiance
    Lp = s.outputs.atmospheric_intrinsic_radiance  # path radiance
    absorb = s.outputs.trans['global_gas'].upward  # absorption transmissivity
    scatter = s.outputs.trans['total_scattering'].upward  # scattering transmissivity
    tau2 = absorb * scatter  # total transmissivity
    # radiance to surface reflectance
    rad = toa_to_rad(bandname)
    ref = rad.subtract(Lp).multiply(math.pi).divide(tau2 * (Edir + Edif))
    return ref

    # all wavebands


output = S2.select('QA60')
for band in ['B1', 'B2', 'B3', 'B4', 'B5', 'B6', 'B7', 'B8', 'B8A', 'B9', 'B10', 'B11', 'B12']:
    print(band)
    output = output.addBands(surface_reflectance(band))



def get_landsat_image():
    # Определение области интереса
    roi = ee.Geometry.Point([108.165, 53.558]).buffer(1000)


    landsat_collection = ee.ImageCollection('LANDSAT/LC08/C01/T1_SR') \
        .filterDate('209-01-01', '2019-12-31') \
        .filterBounds(roi) \
        .sort('CLOUD_COVER')


    image = ee.Image(landsat_collection.first())


    vis_params = {'bands': ['B4', 'B3', 'B2'], 'min': 0, 'max': 3000, 'gamma': 1.4}
    image = image.visualize(**vis_params)

    return image


def func1(img):
    S2 = ee.Image(img)
    date = S2.date()
    # top of atmosphere reflectance
    toa = S2.divide(10000)

    info = S2.getInfo()['properties']
    scene_date = datetime.utcfromtimestamp(
        info['system:time_start'] / 1000)
    solar_z = info['MEAN_SOLAR_ZENITH_ANGLE']

    h2o = S2.water( date).getInfo()
    o3 = S2.ozone( date).getInfo()
    aot = S2.aerosol(date).getInfo()

    SRTM = ee.Image('CGIAR/SRTM90_V4')
    alt = SRTM.reduceRegion(reducer=ee.Reducer.mean(), geometry=toa.centroid()).get('elevation').getInfo()
    km = alt / 1000

    # Instantiate
    s = SixS()

    # Atmospheric constituents
    s.atmos_profile = toa.UserWaterAndOzone(h2o, o3)
    s.aero_profile = toa.Maritime
    s.aot550 = aot

    # Earth-Sun-satellite geometry
    s.geometry = Geometry.User()
    s.geometry.view_z = 0
    s.geometry.solar_z = solar_z  # solar zenith angle
    s.geometry.month = scene_date.month
    s.geometry.day = scene_date.day
    s.altitudes.set_sensor_satellite_level()
    s.altitudes.set_target_custom_altitude(km)

def count_white_pixels(image_path):
    image = Image.open(image_path)
    image = image.convert("RGB")  # Преобразуем изображение в режим RGB

    width, height = image.size
    white_pixel_count = 0

    for y in range(height):
        for x in range(width):
            r, g, b = image.getpixel((x, y))

            if r == 255 and g == 255 and b == 255:
                white_pixel_count += 1

    return white_pixel_count


# Функция для скачивания изображения на диск
def download_image(image_url, file_path):
    response = requests.get(image_url)
    with open(file_path, "wb") as f:
        f.write(response.content)


for ims in iNames:
  # Открываем файл GeoTIFF с помощью библиотеки rasterio
  imag = 'images/roi/' +  ims
  with rasterio.open(imag) as src:

      # Читаем данные из каналов спектра
      green_band = src.read(3).astype('float32')
      red_band = src.read(4).astype('float32')
      blue_band = src.read(2).astype('float32')
      nir_band = src.read(5).astype('float32')

      # Вычисляем индекс MNDWI
      mndwi = (green_band - nir_band) / (green_band + nir_band)

      # Вычисляем индекс EWI
      ewi = 2.5 * ((nir_band - red_band) / (nir_band + 6 * red_band - 7.5 * blue_band + 1))

      # Вычисляем индекс TSS
      tss = np.log10((blue_band + green_band) / (2 * red_band))

      # Вычисляем индекс Chl-a
      chl_a = np.log10(nir_band / red_band) * 0.5 - 0.17

      # Выводим результаты на экран
  indices = [mndwi,ewi,tss,chl_a]


def save_dict_to_excel(dictionary, filename):
    workbook = Workbook()
    sheet = workbook.active

    # Записываем ключи в первую строку таблицы
    keys = list(dictionary.keys())
    sheet.append(keys)

    # Записываем значения словаря в следующие строки
    values = list(dictionary.values())
    sheet.append(values)

    workbook.save(filename)

def load_dict_from_excel(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active

    # Получаем ключи из первой строки таблицы
    keys = [cell.value for cell in sheet[1]]

    # Получаем значения из следующих строк
    values = [cell.value for cell in sheet[2]]

    dictionary = dict(zip(keys, values))
    return dictionary


def load_dict_from_excel(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active

    # Получаем ключи из первой строки таблицы
    keys = [cell.value for cell in sheet[1]]

    # Получаем значения из следующих строк
    values = [cell.value for cell in sheet[2]]

    dictionary = dict(zip(keys, values))
    return dictionary


#  indices = [mndwi,ewi,tss,chl_a]
loaded_ind = load_dict_from_excel('indices.xlsx')
loaded_temp = load_dict_from_excel('temp.xlsx')
answ_ind = [[0,0],[0,0],[0,0],[0,0],[0,0]] # Первое число - 1 -> превышен порог по индексу
                                          # Второе число - 1 -> индекс слишком сильно скакнул
i = 0
prev = current = count_white_pixels('images/roi/' + iNames[0])
for ims in iNames:
    i+=1
    imag = 'images/roi/' +  ims
    if i%2==0 :
        prev = count_white_pixels(imag)
    else:
        current = count_white_pixels(imag)
    if abs(current-prev)>100:
        diff = current-prev



    if loaded_ind[ims][0] <0.4 or loaded_ind[ims][0] >0.6:
        answ_ind[0][0] = 1
    if abs(loaded_ind[ims][0] - iNames[ims][0]) >0.15:
        answ_ind[0][1]=1

    if loaded_ind[ims][1] >0.1:
        answ_ind[1][0] = 1
    if abs(loaded_ind[ims][1] - iNames[ims][1]) >0.15:
        answ_ind[1][1]=1

    if loaded_ind[ims][2] <0.8 or loaded_ind[ims][0] >1.2:
        answ_ind[2][0] = 1
    if abs(loaded_ind[ims][2] - iNames[ims][2]) >0.2:
        answ_ind[2][1]=1

    if loaded_ind[ims][3] <0.1 or loaded_ind[ims][3] >0.5:
        answ_ind[3][0] = 1
    if abs(loaded_ind[ims][3] - iNames[ims][3]) >0.2:
        answ_ind[3][1]=1

answ_temp = [0]
avgTemp = 0
for ims in iNames:
  avgTemp = 0
  for i in range(25):
    avgTemp += loaded_temp[ims][i]
  avgTemp = avgTemp/25
  for i in range(25):
    if abs(loaded_temp[ims][i] - avgTemp) >nTemp:
      answ_temp[0].append(loaded_temp[ims][i] - avgTemp)


save_dict_to_excel(answ_ind, 'answ_ind.xlsx')
save_dict_to_excel(answ_temp, 'answ_temp.xlsx')



def write_to_excel(data):
    # Создаем DataFrame из словаря
    df = pd.DataFrame.from_dict(data, orient='index')
    df = df.transpose()

    # Задаем заголовки столбцов
    df.columns = ['water_body_name', 'mndwi', 'ewi', 'tss', 'chl-a', 'date_image']

    # Записываем DataFrame в Excel-файл
    df.to_excel('results.xlsx', index=False)

write_to_excel(answ_ind)